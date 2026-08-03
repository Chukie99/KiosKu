from datetime import date, datetime, timedelta

from fastapi import APIRouter, Depends, HTTPException
from fastapi.responses import StreamingResponse
from sqlalchemy import func, select
from sqlalchemy.orm import Session, joinedload

from ..database import get_db
from ..models import Product, StockLog, Transaction, TransactionItem

router = APIRouter()


def _summarize(rows: list) -> dict:
    return {
        "total_transactions": len(rows),
        "omzet": round(sum(r.total_amount for r in rows), 2),
        "avg_belanja": round(sum(r.total_amount for r in rows) / len(rows), 2) if rows else 0,
        "items_sold": sum(i.qty for r in rows for i in r.items),
    }


@router.get("/reports/daily")
def daily_report(dt: str | None = None, db: Session = Depends(get_db)):
    day = dt or datetime.now().strftime("%Y-%m-%d")
    rows = db.scalars(
        select(Transaction)
        .where(Transaction.created_at >= f"{day} 00:00:00", Transaction.created_at <= f"{day} 23:59:59")
        .order_by(Transaction.created_at)
    ).all()
    summary = _summarize(rows)
    by_method = {}
    for r in rows:
        key = r.payment_method
        by_method[key] = round(by_method.get(key, 0) + r.total_amount, 2)
    return {"date": day, "summary": summary, "by_method": by_method, "transactions": len(rows)}


@router.get("/reports/monthly")
def monthly_report(month: int | None = None, year: int | None = None, db: Session = Depends(get_db)):
    now = datetime.now()
    m = month or now.month
    y = year or now.year
    start = datetime(y, m, 1)
    end = (start + timedelta(days=32)).replace(day=1) - timedelta(seconds=1)
    rows = db.scalars(
        select(Transaction).where(Transaction.created_at >= start, Transaction.created_at <= end)
    ).all()
    summary = _summarize(rows)
    daily = {}
    for r in rows:
        day_key = r.created_at.strftime("%Y-%m-%d")
        daily[day_key] = round(daily.get(day_key, 0) + r.total_amount, 2)
    return {"month": m, "year": y, "summary": summary, "daily": daily}


@router.get("/reports/top-products")
def top_products(date_from: str | None = None, date_to: str | None = None, limit: int = 10, db: Session = Depends(get_db)):
    q = (
        select(
            TransactionItem.product_id,
            TransactionItem.product_name_snapshot,
            func.sum(TransactionItem.qty).label("qty_sold"),
            func.sum(TransactionItem.subtotal).label("revenue"),
        )
        .join(Transaction, Transaction.id == TransactionItem.transaction_id)
        .where(Transaction.status == "selesai")
    )
    if date_from:
        q = q.where(Transaction.created_at >= f"{date_from} 00:00:00")
    if date_to:
        q = q.where(Transaction.created_at <= f"{date_to} 23:59:59")
    q = q.group_by(TransactionItem.product_id, TransactionItem.product_name_snapshot).order_by(func.sum(TransactionItem.qty).desc()).limit(limit)
    return [
        {
            "product_id": pid,
            "product_name": name,
            "qty_sold": qty,
            "revenue": round(rev, 2),
        }
        for pid, name, qty, rev in db.execute(q).all()
    ]


@router.get("/reports/summary")
def reports_summary(db: Session = Depends(get_db)):
    today = datetime.now().strftime("%Y-%m-%d")
    monthly_start = datetime.now().replace(day=1)
    today_rows = db.scalars(select(Transaction).where(Transaction.created_at >= f"{today} 00:00:00")).all()
    month_rows = db.scalars(select(Transaction).where(Transaction.created_at >= monthly_start)).all()
    low_stock = db.scalars(
        select(Product).where(Product.is_active.is_(True), Product.stock <= Product.stock_alert_threshold)
    ).all()
    return {
        "today": _summarize(today_rows),
        "this_month": _summarize(month_rows),
        "low_stock_count": len(low_stock),
    }


def _generate_xlsx(date_from: str, date_to: str, db: Session) -> bytes:
    from io import BytesIO

    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill

    rows = db.scalars(
        select(Transaction)
        .options(joinedload(Transaction.items))
        .where(Transaction.created_at >= f"{date_from} 00:00:00", Transaction.created_at <= f"{date_to} 23:59:59")
    ).unique().all()

    wb = Workbook()
    ws = wb.active
    ws.title = "Transaksi"
    headers = ["No", "Invoice", "Tanggal", "Metode", "Total", "Status"]
    ws.append(headers)
    for c in ws[1]:
        c.font = Font(bold=True)
        c.fill = PatternFill("solid", fgColor="0D6E6E")
        c.font = Font(bold=True, color="FFFFFF")
    for i, t in enumerate(rows, 1):
        ws.append([i, t.invoice_no, t.created_at.strftime("%Y-%m-%d %H:%M"), t.payment_method, t.total_amount, t.status])

    ws2 = wb.create_sheet("Detail Item")
    ws2.append(["Invoice", "Produk", "Satuan", "Qty", "Harga", "Subtotal"])
    for c in ws2[1]:
        c.fill = PatternFill("solid", fgColor="0D6E6E")
        c.font = Font(bold=True, color="FFFFFF")
    for t in rows:
        for item in t.items:
            ws2.append([t.invoice_no, item.product_name_snapshot, item.unit_name, item.qty, item.price_per_unit, item.subtotal])

    buffer = BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


def _generate_pdf(date_from: str, date_to: str, db: Session) -> bytes:
    from io import BytesIO

    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.units import mm
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet

    rows = db.scalars(
        select(Transaction)
        .options(joinedload(Transaction.items))
        .where(Transaction.created_at >= f"{date_from} 00:00:00", Transaction.created_at <= f"{date_to} 23:59:59")
    ).unique().all()

    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4, rightMargin=15 * mm, leftMargin=15 * mm, topMargin=15 * mm, bottomMargin=15 * mm)
    styles = getSampleStyleSheet()
    story = [Paragraph(f"Laporan Penjualan KiosKu", styles["Title"]),
             Paragraph(f"Periode: {date_from} s/d {date_to} ({len(rows)} transaksi)", styles["Normal"]), Spacer(1, 6 * mm)]

    data = [["Invoice", "Tanggal", "Metode", "Total"]]
    for t in rows:
        data.append([t.invoice_no, t.created_at.strftime("%Y-%m-%d %H:%M"), t.payment_method, f"{t.total_amount:,.0f}"])
    table = Table(data, colWidths=[45 * mm, 40 * mm, 35 * mm, 30 * mm])
    table.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#0D6E6E")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
                ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F8FAFC")]),
            ]
        )
    )
    story.append(table)
    doc.build(story)
    return buffer.getvalue()


@router.get("/reports/export")
def export_report(format: str = "xlsx", date_from: str | None = None, date_to: str | None = None, db: Session = Depends(get_db)):
    if format not in ("xlsx", "pdf"):
        raise HTTPException(status_code=422, detail="Format harus xlsx atau pdf")
    date_to = date_to or datetime.now().strftime("%Y-%m-%d")
    date_from = date_from or (datetime.now() - timedelta(days=30)).strftime("%Y-%m-%d")
    if format == "xlsx":
        content = _generate_xlsx(date_from, date_to, db)
        media = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        filename = f"laporan_{date_from}_to_{date_to}.xlsx"
    else:
        content = _generate_pdf(date_from, date_to, db)
        media = "application/pdf"
        filename = f"laporan_{date_from}_to_{date_to}.pdf"
    return StreamingResponse(
        iter([content]),
        media_type=media,
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )
